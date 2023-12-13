from pathlib import Path
from collections import defaultdict

import openpyxl
from PyPDF2 import PdfReader, PdfWriter


def load_xls(filename):
    work_book = openpyxl.load_workbook(filename)
    work_sheet = work_book.active
    return work_book, work_sheet


def write_to_xls(work_sheet, work_book, data, output_xls):
    work_sheet.delete_rows(1, work_sheet.max_row)
    for value in data.values():
        for i in value:
            work_sheet.append(i)

    work_book.save(output_xls)


def clear_and_save(work_sheet, work_book, output_xls):
    work_sheet.delete_cols(7, work_sheet.max_column)
    work_sheet[f'B{work_sheet.max_row + 1}'] = f'=sum(B1:B{work_sheet.max_row})'
    work_book.save(output_xls)


def del_columns(work_sheet):
    work_sheet.delete_cols(15, 6)
    work_sheet.delete_cols(13)
    work_sheet.delete_cols(10)
    work_sheet.delete_cols(8)
    work_sheet.delete_cols(3, 4)
    work_sheet.delete_cols(1)


def preparation(orders_filename, output_xls):
    work_book, work_sheet = load_xls(orders_filename)
    max_row = work_sheet.max_row
    number = 0

    del_columns(work_sheet)

    for i in range(1, max_row + 1):
        if work_sheet[f'A{i}'].value != work_sheet[f'A{i + 1}'].value:
            work_sheet[f'G{i}'] = number
            number += 1

    work_book.save(output_xls)


def sorting(work_sheet):
    dictionary = defaultdict(list)

    for row in work_sheet.iter_rows(min_row=0, values_only=True):
        dictionary[row[0]].append(row)

    sorted_dict = dict(sorted(dictionary.items(), key=lambda x: (len(x[1]), x[1][0][2])))

    return sorted_dict


def pdf(work_sheet, input_filename, output_filename):
    input_pdf = PdfReader(input_filename)
    pdf_pages = []
    sticker_numbers = []

    for i in range(1, work_sheet.max_row + 1):
        cell_value = work_sheet[f'G{i}'].value
        if cell_value is not None:
            sticker_numbers.append(cell_value)

    for number in sticker_numbers:
        pdf_page = input_pdf.pages[number]
        pdf_pages.append(pdf_page)

    pdf_writer = PdfWriter()
    for page in pdf_pages:
        pdf_writer.add_page(page)

    with Path(output_filename).open(mode="wb") as file:
        pdf_writer.write(file)


def main():
    current_path = Path.cwd()
    input_xls = Path(current_path, 'orders.xlsx')
    input_pdf = Path(current_path, 'stickers.pdf')
    output_xls = Path(current_path, 'final.xlsx')
    output_pdf = Path(current_path, 'final.pdf')

    preparation(input_xls, output_xls)

    work_book, work_sheet = load_xls(output_xls)

    sorted_data = sorting(work_sheet)

    write_to_xls(work_sheet, work_book, sorted_data, output_xls)
    pdf(work_sheet, input_pdf, output_pdf)
    clear_and_save(work_sheet, work_book, output_xls)


if __name__ == '__main__':
    main()
